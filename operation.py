from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart,
    BarChart,
    Reference
)

import os
import datetime

# Declaring file path as global variable
file_path = "workbook\\FinanceTracker.xlsx"

# Main worksheet column designations
date_day_column = "B"
date_month_year = "C"
day_column = "D"
time_column = "E"
category_column = "F"
expense_column = "G"

# Sheet names
Main = "Personal Finance"
Category_Data = "Expenses (Category)"
Time_Chart = "Expenses (Monthly)"


# Creates a worksheet in workbook
def create_worksheet(workbook, name: str, position: int, colour: str):
    worksheet = workbook.create_sheet(name, position)
    worksheet.sheet_properties.tabColor = colour
    return worksheet


# Enter data into specific cell in a specific sheet
def enter_data(sheet, cell, data):
    sheet[cell] = data
    return


def create_file():
    # file_check returns True if FinanceTracker.xlsx exists
    file_check = os.path.isfile(file_path)

    # If file exists, pass, otherwise, creates file
    if file_check:
        pass
    else:
        data_book = Workbook()

        # Creates and configures the main worksheet
        worksheet_1 = create_worksheet(data_book, Main, 0, "05CEBD")
        enter_data(worksheet_1, f"{date_day_column}2", "Date - Day")
        enter_data(worksheet_1, f"{date_month_year}2", "Date - Month/Year")
        enter_data(worksheet_1, f"{day_column}2", "Day")
        enter_data(worksheet_1, f"{time_column}2", "Time")
        enter_data(worksheet_1, f"{category_column}2", "Category")
        enter_data(worksheet_1, f"{expense_column}2", "Expense")

        # Creates additional worksheet where chart references are stored
        worksheet_2 = create_worksheet(data_book, Category_Data, 1, "05CEBD")
        enter_data(worksheet_2, f"B2", "Category")
        enter_data(worksheet_2, f"C2", "Expense")

        worksheet_3 = create_worksheet(data_book, Time_Chart, 2, "05CEBD")
        enter_data(worksheet_3, f"B2", "Month/Year")
        enter_data(worksheet_3, f"C2", "Expense")
        data_book.save(file_path)
    return


def open_file():
    os.system(f"start EXCEL.EXE {file_path}")
    return


def get_date_time():
    date_time = datetime.datetime.now()
    day_of_month = date_time.strftime("%d")
    month_year = date_time.strftime("%b-%Y")
    day = date_time.strftime("%A")
    time = date_time.strftime("%H:%M:%S %p")
    return day_of_month, month_year, day, time


def load_book(sheet):
    wb = load_workbook(file_path)
    ws = wb[sheet]
    try:
        max_row = str(ws.max_row + 1)
    except AttributeError:
        max_row = None

    return wb, ws, max_row


def user_input_data(sheet, category, amount):
    wb, ws, max_row = load_book(sheet)

    date_day, date_my, day, time = get_date_time()

    enter_data(ws, f"{date_day_column}{max_row}", date_day)  # Entering day of date
    enter_data(ws, f"{date_month_year}{max_row}", date_my)  # Entering month and year
    enter_data(ws, f"{day_column}{max_row}", day)  # Entering day
    enter_data(ws, f"{time_column}{max_row}", time)  # Entering time
    enter_data(ws, f"{category_column}{max_row}", category)  # Entering category
    enter_data(ws, f"{expense_column}{max_row}", amount)  # Entering expense

    wb.save(file_path)
    return


def filter_data(sheet, column_name, column_number):
    filtered_list = []
    ignored = [None, column_name]

    for row in sheet.iter_rows(max_col=column_number, min_col=column_number):
        for cell in row:
            if cell.value not in filtered_list and cell.value not in ignored:
                filtered_list.append(cell.value)

    return filtered_list


def load_chart_data(sheet, column_name, data_sheet):
    chart_data = {}

    if column_name == "Category":
        column_number = 6  # Returns category column number
        expense_number = 1
    else:
        column_number = 3  # Returns month/year column number
        expense_number = 4

    item_set = filter_data(sheet, column_name, column_number)

    for item in item_set:
        item_data = []

        for row in sheet.iter_rows(min_row=3, min_col=column_number, max_col=7):
            if row[0].value == item:
                item_data.append(float(row[expense_number].value))

        total = sum(i for i in item_data)
        chart_data[item] = total

    r = 3
    for item in chart_data:
        data_sheet.cell(row=r, column=2).value = item
        data_sheet.cell(row=r, column=3).value = chart_data[item]
        r += 1

    return r


def create_pie_chart():
    wb, ws, max_row = load_book(Main)
    chart_worksheet = wb[Category_Data]
    cs = wb.create_chartsheet("Category Chart")

    r = load_chart_data(ws, "Category", chart_worksheet)

    category_expenses = PieChart()
    labels = Reference(chart_worksheet, min_col=2, min_row=3, max_row = r-1)
    data = Reference(chart_worksheet, min_col=3, min_row=2, max_row = r-1)
    category_expenses.add_data(data, titles_from_data=True)
    category_expenses.set_categories(labels)
    category_expenses.title = "Expenses by Category"

    cs.add_chart(category_expenses)
    wb.save(file_path)


def create_bar_chart():
    wb, ws, max_row = load_book(Main)
    cs = wb.create_chartsheet("Time Chart")
    chart_worksheet = wb[Time_Chart]

    r = load_chart_data(ws, "Date - Month/Year", chart_worksheet)

    chart = BarChart()
    chart.title = "Expenses against Time"
    chart.style = 13
    chart.x_axis.title = "Time"
    chart.y_axis.title = "Expenses"

    data = Reference(chart_worksheet, min_col=3, min_row=3, max_row=r-1)
    chart.add_data(data)

    dates = Reference(chart_worksheet, min_col=2, min_row=3, max_row=r-1)
    chart.set_categories(dates)

    cs.add_chart(chart)
    wb.save(file_path)


def auto_adjust_column():
    # Load workbook and get all sheet names
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames

    # Iterates over all sheets in workbook
    for sheet in sheet_list:
        ws = wb[sheet]
        try:
            for column in ws.columns:
                max_length = 0
                column_name = column[0].column_letter

                for row in column:  # Finding the max width of each column
                    if row.value is not None:
                        if len(str(row.value)) > max_length:
                            max_length = len(row.value)
                new_width = (max_length + 3) * 1.3  # Setting new column width based on max width

                ws.column_dimensions[column_name].width = new_width
        except AttributeError as e:    # Ignores chart sheets
            pass
    wb.save(file_path)
    return


def delete_sheet(sheet):
    try:
        wb, ws, max_row = load_book(sheet)
        wb.remove(ws)
        wb.save(file_path)
    except KeyError:
        pass
    return


def initialize_file():
    create_file()
    delete_sheet("Sheet")


def open_function():
    delete_sheet("Category Chart")
    delete_sheet("Time Chart")

    create_bar_chart()
    create_pie_chart()
    auto_adjust_column()
    open_file()